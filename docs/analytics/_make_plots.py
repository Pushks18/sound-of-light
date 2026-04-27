"""
Generate analytics plots from /Users/praveen/Downloads/alter-ego beta.xlsx.
Outputs PNGs into the same directory as this script.
"""
import os
from collections import Counter, defaultdict
from datetime import datetime
import statistics

import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
SRC = "/Users/praveen/Downloads/alter-ego beta.xlsx"

DARK_BG = "#1a1a1f"
PANEL = "#22222a"
FG = "#e6e6f0"
GRID = "#3a3a44"
ACC_SLASH = "#ffd24a"
ACC_BULLET = "#9bc1ff"
ACC_DASH = "#9affc8"
ACC_TRAP = "#ff8a8a"
ACC_UNKNOWN = "#9a9aaa"
ACC_NEUTRAL = "#cdd1d8"
ACC_HIGHLIGHT = "#ff7c7c"

plt.rcParams.update({
    "figure.facecolor": DARK_BG,
    "axes.facecolor": PANEL,
    "axes.edgecolor": FG,
    "axes.labelcolor": FG,
    "axes.titlecolor": FG,
    "xtick.color": FG,
    "ytick.color": FG,
    "text.color": FG,
    "axes.grid": True,
    "grid.color": GRID,
    "grid.alpha": 0.6,
    "axes.spines.top": False,
    "axes.spines.right": False,
    "font.family": "DejaVu Sans",
    "font.size": 11,
})


def load():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb["Form Responses 1"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers) if h is not None}
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([c.value for c in ws[r]])
    return idx, rows


idx, rows = load()


def get(row, name):
    i = idx.get(name)
    if i is None:
        return None
    return row[i]


# ─────────────────────────────────────────────────────────────────────────────
# Plot 1: Per-day kill method mix (stacked %)
# ─────────────────────────────────────────────────────────────────────────────
def plot_per_day_method_mix():
    day_method = defaultdict(Counter)
    for row in rows:
        ts = row[0]
        if not ts or get(row, "run_outcome") is None:
            continue
        d = ts.date()
        for m in ("slash", "bullet", "dash", "trap", "unknown"):
            day_method[d][m] += get(row, m + "_kills") or 0

    days = sorted(day_method)
    methods = ["slash", "bullet", "dash", "trap", "unknown"]
    colors = [ACC_SLASH, ACC_BULLET, ACC_DASH, ACC_TRAP, ACC_UNKNOWN]

    pct = {m: [] for m in methods}
    totals = []
    for d in days:
        total = sum(day_method[d].values()) or 1
        totals.append(total)
        for m in methods:
            pct[m].append(day_method[d][m] / total * 100)

    fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(11, 7.5),
                                    gridspec_kw={"height_ratios": [3, 1]})
    bottom = np.zeros(len(days))
    for m, c in zip(methods, colors):
        vals = np.array(pct[m])
        ax1.bar(range(len(days)), vals, bottom=bottom, label=m, color=c,
                edgecolor=DARK_BG, linewidth=1.0)
        bottom += vals
    ax1.set_ylim(0, 100)
    ax1.set_ylabel("share of kills (%)")
    ax1.set_title("Kill-method mix by beta day  (n=586 kills across 48 runs)",
                  fontsize=13, pad=14)
    ax1.set_xticks(range(len(days)))
    ax1.set_xticklabels([d.isoformat() for d in days], rotation=30, ha="right")
    ax1.legend(loc="lower center", bbox_to_anchor=(0.5, -0.5), ncol=5,
               frameon=False)

    ax2.bar(range(len(days)), totals, color=ACC_NEUTRAL,
            edgecolor=DARK_BG, linewidth=1.0)
    ax2.set_ylabel("total kills")
    ax2.set_xticks(range(len(days)))
    ax2.set_xticklabels([d.isoformat() for d in days], rotation=30, ha="right")
    for x, t in enumerate(totals):
        ax2.text(x, t, f"{int(t)}", ha="center", va="bottom",
                 color=FG, fontsize=9)

    plt.tight_layout()
    out = os.path.join(OUT_DIR, "01_kill_method_mix_by_day.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


# ─────────────────────────────────────────────────────────────────────────────
# Plot 2: TTK histogram by kill method
# ─────────────────────────────────────────────────────────────────────────────
def plot_ttk_by_method():
    by_method = defaultdict(list)
    for row in rows:
        if get(row, "ttk_seconds") is None:
            continue
        en = get(row, "ttk_enemy_name") or ""
        if "Clone" in en:
            continue
        m = get(row, "ttk_kill_method")
        s = get(row, "ttk_seconds")
        if m and s is not None:
            by_method[m].append(s)

    fig, ax = plt.subplots(figsize=(10.5, 5.5))
    methods = [("slash", ACC_SLASH), ("bullet", ACC_BULLET),
               ("dash", ACC_DASH), ("trap", ACC_TRAP)]

    bins = np.linspace(0, 8, 33)
    for m, c in methods:
        data = [d for d in by_method.get(m, []) if d <= 8]
        if not data:
            continue
        ax.hist(data, bins=bins, color=c, alpha=0.65, label=f"{m} (n={len(by_method[m])})",
                edgecolor=DARK_BG, linewidth=0.6)

    for m, c in methods:
        if by_method.get(m):
            med = statistics.median(by_method[m])
            ax.axvline(med, color=c, linestyle="--", alpha=0.95, linewidth=1.6)
            ax.text(med, ax.get_ylim()[1] * 0.95 if ax.get_ylim()[1] else 1,
                    f"  {m} median {med:.2f}s",
                    color=c, fontsize=9.5, va="top", rotation=90)

    ax.set_xlabel("time-to-kill (seconds)")
    ax.set_ylabel("kill count")
    ax.set_xlim(0, 8)
    ax.set_title("Time-to-kill distribution by player kill method  (excluding Crimson clones)",
                 fontsize=13, pad=14)
    ax.legend(loc="upper right", frameon=False)
    plt.tight_layout()
    out = os.path.join(OUT_DIR, "02_ttk_by_method.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


# ─────────────────────────────────────────────────────────────────────────────
# Plot 3: Mean-vs-median TTK gap per enemy spawn slot
# ─────────────────────────────────────────────────────────────────────────────
def plot_mean_median_gap():
    per_enemy = defaultdict(list)
    for row in rows:
        if get(row, "ttk_seconds") is None:
            continue
        en = get(row, "ttk_enemy_name") or ""
        if "Clone" in en or "Skitter" in en:
            continue
        per_enemy[en].append(get(row, "ttk_seconds"))

    rows_data = []
    for en, durs in per_enemy.items():
        if len(durs) < 15:
            continue
        rows_data.append((en, len(durs), statistics.median(durs),
                          statistics.mean(durs)))
    rows_data.sort(key=lambda x: -x[1])  # by sample count desc

    enemies = [r[0] for r in rows_data]
    medians = [r[2] for r in rows_data]
    means = [r[3] for r in rows_data]
    counts = [r[1] for r in rows_data]

    fig, ax = plt.subplots(figsize=(11, 6.5))
    y = np.arange(len(enemies))

    for yi, med, mn in zip(y, medians, means):
        ax.plot([med, mn], [yi, yi], color=ACC_NEUTRAL, alpha=0.6,
                linewidth=2.5, zorder=1)
    ax.scatter(medians, y, color=ACC_BULLET, s=85, zorder=3,
               label="median TTK", edgecolor=DARK_BG, linewidth=1.0)
    ax.scatter(means, y, color=ACC_HIGHLIGHT, s=85, zorder=3,
               label="mean TTK", edgecolor=DARK_BG, linewidth=1.0)

    for yi, mn, n in zip(y, means, counts):
        ax.text(mn + 0.05, yi, f"  n={n}", color=FG, va="center", fontsize=9)

    ax.set_yticks(y)
    ax.set_yticklabels(enemies)
    ax.set_xlabel("seconds")
    ax.set_title("TTK mean vs median per enemy spawn slot — long tail visible everywhere\n"
                 "median = typical 1-slash kill, mean pulled right by stuck/fled survivors",
                 fontsize=12.5, pad=14)
    ax.legend(loc="lower right", frameon=False)
    ax.invert_yaxis()
    plt.tight_layout()
    out = os.path.join(OUT_DIR, "03_ttk_mean_vs_median.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


# ─────────────────────────────────────────────────────────────────────────────
# Plot 4: Damage interaction histogram + fatal-vs-survived overlay
# ─────────────────────────────────────────────────────────────────────────────
def plot_interaction_damage():
    fatal = []
    survived = []
    for row in rows:
        if get(row, "interaction_duration_seconds") is None:
            continue
        dmg = get(row, "interaction_damage_taken") or 0
        end = get(row, "interaction_end_reason")
        if end == "death":
            fatal.append(dmg)
        else:
            survived.append(dmg)

    bins = np.arange(0.5, 10.5, 1)
    fig, ax = plt.subplots(figsize=(10.5, 5.5))
    ax.hist(survived, bins=bins, color=ACC_BULLET, alpha=0.85,
            label=f"survived chain (n={len(survived)}, mean {statistics.mean(survived):.2f} dmg)",
            edgecolor=DARK_BG, linewidth=0.8)
    ax.hist(fatal, bins=bins, color=ACC_HIGHLIGHT, alpha=0.85,
            label=f"fatal chain (n={len(fatal)}, mean {statistics.mean(fatal):.2f} dmg)",
            edgecolor=DARK_BG, linewidth=0.8)

    ax.axvline(3, color=ACC_TRAP, linestyle=":", linewidth=1.5)
    ax.text(3, ax.get_ylim()[1] * 0.92, "  player base HP = 3", color=ACC_TRAP,
            fontsize=10)

    ax.set_xticks(np.arange(1, 10))
    ax.set_xlabel("damage taken in a single interaction (chain)")
    ax.set_ylabel("interaction count")
    ax.set_title("Damage interaction histogram — fatal chains are hit-stacked, not longer\n"
                 "(1,000 chains • 89.4% timeout / 10.6% death)",
                 fontsize=13, pad=14)
    ax.legend(loc="upper right", frameon=False)
    plt.tight_layout()
    out = os.path.join(OUT_DIR, "04_damage_chain_distribution.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


# ─────────────────────────────────────────────────────────────────────────────
# Plot 5: Per-run resource economy collapse (zero-method runs)
# ─────────────────────────────────────────────────────────────────────────────
def plot_per_run_method_usage():
    runs_data = []  # [(slash%, bullet%, dash%, trap%)]
    pure_slash = 0
    zero_bullet = 0
    zero_dash = 0
    zero_trap = 0
    n = 0
    for row in rows:
        if get(row, "run_outcome") is None:
            continue
        sk = get(row, "slash_kills") or 0
        bk = get(row, "bullet_kills") or 0
        dk = get(row, "dash_kills") or 0
        trk = get(row, "trap_kills") or 0
        tk = sk + bk + dk + trk
        if tk == 0:
            continue  # exclude 0-kill runs (player died with no kills)
        runs_data.append((sk / tk * 100, bk / tk * 100, dk / tk * 100, trk / tk * 100))
        n += 1
        if sk == tk:
            pure_slash += 1
        if bk == 0:
            zero_bullet += 1
        if dk == 0:
            zero_dash += 1
        if trk == 0:
            zero_trap += 1

    runs_data.sort(key=lambda r: -r[0])
    arr = np.array(runs_data)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5.6),
                                    gridspec_kw={"width_ratios": [2.2, 1]})

    bottom = np.zeros(len(runs_data))
    for col, lbl, c in zip(range(4),
                           ["slash", "bullet", "dash", "trap"],
                           [ACC_SLASH, ACC_BULLET, ACC_DASH, ACC_TRAP]):
        vals = arr[:, col]
        ax1.bar(range(len(runs_data)), vals, bottom=bottom, color=c, label=lbl,
                edgecolor=DARK_BG, linewidth=0.4, width=1.0)
        bottom += vals
    ax1.set_xlim(-0.5, len(runs_data) - 0.5)
    ax1.set_ylim(0, 100)
    ax1.set_xlabel(f"individual runs (sorted by slash share, n={n} non-zero-kill runs)")
    ax1.set_ylabel("share of kills (%)")
    ax1.set_title("Per-run kill-method composition — most runs are slash-dominant",
                  fontsize=12.5, pad=14)
    ax1.legend(loc="lower left", frameon=False, ncol=4)

    cats = ["pure-slash\nruns", "runs that\nnever shoot", "runs that\nnever dash",
            "runs that\nnever trap"]
    pcts = [pure_slash / n * 100, zero_bullet / n * 100,
            zero_dash / n * 100, zero_trap / n * 100]
    bars = ax2.bar(cats, pcts, color=[ACC_SLASH, ACC_BULLET, ACC_DASH, ACC_TRAP],
                   edgecolor=DARK_BG, linewidth=1.0)
    ax2.set_ylim(0, 100)
    ax2.set_ylabel("% of runs")
    ax2.set_title("Resource-economy collapse",
                  fontsize=12.5, pad=14)
    for b, p in zip(bars, pcts):
        ax2.text(b.get_x() + b.get_width() / 2, b.get_height() + 2,
                 f"{p:.0f}%", ha="center", color=FG, fontsize=11)
    ax2.set_xticklabels(cats, fontsize=9.5)

    plt.tight_layout()
    out = os.path.join(OUT_DIR, "05_per_run_resource_collapse.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


# ─────────────────────────────────────────────────────────────────────────────
# Plot 6: Skitter vs basic enemies — method mix shift
# ─────────────────────────────────────────────────────────────────────────────
def plot_skitter_vs_basic():
    basic = Counter()
    skit = Counter()
    skit_ttks = []
    basic_ttks = []
    for row in rows:
        if get(row, "ttk_seconds") is None:
            continue
        en = get(row, "ttk_enemy_name") or ""
        m = get(row, "ttk_kill_method")
        s = get(row, "ttk_seconds")
        if "Clone" in en:
            continue
        if "Skitter" in en:
            skit[m] += 1
            skit_ttks.append(s)
        else:
            basic[m] += 1
            basic_ttks.append(s)

    methods = ["slash", "bullet", "dash", "trap"]
    colors = [ACC_SLASH, ACC_BULLET, ACC_DASH, ACC_TRAP]

    def shares(c):
        t = sum(c.values()) or 1
        return [c[m] / t * 100 for m in methods]

    basic_pct = shares(basic)
    skit_pct = shares(skit)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12.5, 5.5))

    x = np.arange(len(methods))
    w = 0.36
    ax1.bar(x - w / 2, basic_pct, w, color=[ACC_NEUTRAL] * len(methods),
            label=f"basic enemies (n={sum(basic.values())})",
            edgecolor=DARK_BG, linewidth=1.0)
    bars2 = ax1.bar(x + w / 2, skit_pct, w, color=colors,
                    label=f"Skitter (n={sum(skit.values())})",
                    edgecolor=DARK_BG, linewidth=1.0)
    for xi, (a, b) in enumerate(zip(basic_pct, skit_pct)):
        delta = b - a
        ax1.text(xi + w / 2, b + 2, f"{b:.0f}%", ha="center", color=FG,
                 fontsize=10)
        ax1.text(xi - w / 2, a + 2, f"{a:.0f}%", ha="center", color=FG,
                 fontsize=10)
        if abs(delta) >= 5:
            color = ACC_HIGHLIGHT if delta < 0 else "#9affc8"
            ax1.text(xi, max(a, b) + 8, f"Δ{delta:+.0f}pp", ha="center",
                     color=color, fontsize=11, fontweight="bold")
    ax1.set_xticks(x)
    ax1.set_xticklabels(methods)
    ax1.set_ylabel("share of kills (%)")
    ax1.set_ylim(0, 100)
    ax1.set_title("Skitter shifts the kit — bullet share doubles",
                  fontsize=12.5, pad=14)
    ax1.legend(loc="upper right", frameon=False)

    bins = np.linspace(0, 12, 25)
    ax2.hist(basic_ttks, bins=bins, color=ACC_NEUTRAL, alpha=0.85,
             label=f"basic (median {statistics.median(basic_ttks):.2f}s, mean {statistics.mean(basic_ttks):.2f}s)",
             edgecolor=DARK_BG, linewidth=0.6)
    ax2.hist(skit_ttks, bins=bins, color=ACC_HIGHLIGHT, alpha=0.75,
             label=f"Skitter (median {statistics.median(skit_ttks):.2f}s, mean {statistics.mean(skit_ttks):.2f}s)",
             edgecolor=DARK_BG, linewidth=0.6)
    ax2.set_xlabel("time-to-kill (seconds)")
    ax2.set_ylabel("count")
    ax2.set_title("Skitter survives engagement 3-4× longer",
                  fontsize=12.5, pad=14)
    ax2.legend(loc="upper right", frameon=False)
    ax2.set_xlim(0, 12)

    plt.tight_layout()
    out = os.path.join(OUT_DIR, "06_skitter_vs_basic.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


# ─────────────────────────────────────────────────────────────────────────────
# Plot 7: Cumulative timeline — events per day with change markers
# ─────────────────────────────────────────────────────────────────────────────
def plot_timeline():
    by_day_event = Counter()
    for row in rows:
        ts = row[0]
        if not ts:
            continue
        by_day_event[ts.date()] += 1
    days = sorted(by_day_event)
    counts = [by_day_event[d] for d in days]
    cum = np.cumsum(counts)

    fig, ax = plt.subplots(figsize=(11.5, 5.5))
    ax.fill_between(range(len(days)), cum, color=ACC_BULLET, alpha=0.25)
    ax.plot(range(len(days)), cum, color=ACC_BULLET, linewidth=2.4,
            marker="o", markersize=5)
    for x, (d, c) in enumerate(zip(days, counts)):
        if c >= 100:
            ax.annotate(f"{c}", xy=(x, cum[x]), xytext=(0, 6),
                        textcoords="offset points", ha="center",
                        color=ACC_NEUTRAL, fontsize=9)

    # Mark commits that responded to the data
    iso = [d.isoformat() for d in days]

    def mark(date_str, label, y_offset=0.85, color=ACC_HIGHLIGHT):
        if date_str in iso:
            xi = iso.index(date_str)
            ax.axvline(xi, color=color, linestyle="--", alpha=0.7, linewidth=1.4)
            ax.text(xi, max(cum) * y_offset, label, color=color,
                    fontsize=9.5, rotation=90, va="top", ha="right")

    mark("2026-04-21", "HP packs + hunting fix", 0.55, "#9affc8")
    mark("2026-04-25", "Skitter slash resistance,\nclose-range, proximity,\nclone phase", 0.95, "#9affc8")
    mark("2026-04-26", "Room-clear rewards", 0.40, "#9affc8")

    ax.set_xticks(range(len(days)))
    ax.set_xticklabels([d.isoformat() for d in days], rotation=30, ha="right",
                       fontsize=9)
    ax.set_ylabel("cumulative telemetry events")
    ax.set_title("Beta telemetry timeline — green dashes mark analytics-driven commits",
                 fontsize=13, pad=14)
    plt.tight_layout()
    out = os.path.join(OUT_DIR, "07_timeline.png")
    plt.savefig(out, dpi=140, bbox_inches="tight", facecolor=DARK_BG)
    plt.close(fig)
    print("wrote", out)


if __name__ == "__main__":
    plot_per_day_method_mix()
    plot_ttk_by_method()
    plot_mean_median_gap()
    plot_interaction_damage()
    plot_per_run_method_usage()
    plot_skitter_vs_basic()
    plot_timeline()
    print("\nAll plots written to", OUT_DIR)
